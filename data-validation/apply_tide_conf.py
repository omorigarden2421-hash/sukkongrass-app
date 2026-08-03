# -*- coding: utf-8 -*-
"""耐潮性(42/43/44)＋信頼度(29)のWeb再検証結果をマスタ2ファイルに適用し修正ログを出力"""
import openpyxl, json, csv, re, sys

FILES = [
    '納品_最終版/宿根草マスタ_検証済み_最終版.xlsx',
    '納品_最終版/宿根草マスタ_調査完了項目_赤塗り_最終版.xlsx',
]

# 結果集約
res = {}
for i in range(6):
    for r in json.load(open(f'_batch/result{i}.json')):
        res[int(r['row'])] = r

# 対象種別（耐潮性 / 信頼度）
need = {}
for r in csv.DictReader(open('耐潮性_信頼度_再検証対象リスト.csv', encoding='utf-8-sig')):
    need[int(r['excel_row'])] = r['対象']

COL_TIDE, COL_SRC, COL_REASON, COL_CONF, COL_NAME = 42, 43, 44, 29, 3

# 検証済み版で行対応を確認しつつログ生成（1回だけ）
wb0 = openpyxl.load_workbook(FILES[0])
ws0 = wb0.active
log_rows = []
mismatch = []
for row, r in sorted(res.items()):
    name_cell = str(ws0.cell(row=row, column=COL_NAME).value or '').replace('\n', ' ')
    old_tide = str(ws0.cell(row=row, column=COL_TIDE).value or '').strip()
    old_src  = str(ws0.cell(row=row, column=COL_SRC).value or '').strip()
    old_conf = str(ws0.cell(row=row, column=COL_CONF).value or '').strip()
    new_tide = str(r.get('ver_tide', '')).strip()
    new_src  = str(r.get('source', '')).strip()
    new_reason = str(r.get('reason', '')).strip()
    new_conf = str(r.get('confidence', '')).strip()
    kinds = need.get(row, '耐潮性')
    log_rows.append({
        'excel_row': row, 'plant': name_cell, '対象': kinds,
        '耐潮性_旧': old_tide, '耐潮性_新': new_tide,
        '信頼度_旧': old_conf, '信頼度_新': (new_conf if '信頼度' in kinds else old_conf),
        'source': new_src, 'reason': new_reason,
    })

with open('耐潮性_信頼度_再検証_修正ログ.csv', 'w', encoding='utf-8-sig', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['excel_row','plant','対象','耐潮性_旧','耐潮性_新','信頼度_旧','信頼度_新','source','reason'])
    w.writeheader(); w.writerows(log_rows)

tide_chg = sum(1 for x in log_rows if x['耐潮性_旧'] != x['耐潮性_新'])
conf_fill = sum(1 for x in log_rows if '信頼度' in x['対象'])
print(f'ログ出力: {len(log_rows)}行  耐潮性変更 {tide_chg}  信頼度充填 {conf_fill}')

# 両ファイルに適用
for path in FILES:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    n = 0
    for row, r in res.items():
        kinds = need.get(row, '耐潮性')
        # 名前照合（ズレ検知）
        nm = str(ws.cell(row=row, column=COL_NAME).value or '')
        if not nm.strip():
            mismatch.append((path, row)); continue
        ws.cell(row=row, column=COL_TIDE).value  = str(r.get('ver_tide','')).strip()
        ws.cell(row=row, column=COL_SRC).value   = str(r.get('source','')).strip()
        ws.cell(row=row, column=COL_REASON).value = str(r.get('reason','')).strip()
        if '信頼度' in kinds:
            ws.cell(row=row, column=COL_CONF).value = str(r.get('confidence','')).strip()
        n += 1
    wb.save(path)
    print(f'適用完了: {path}  {n}行更新')

if mismatch:
    print('WARN 空名行:', mismatch)
