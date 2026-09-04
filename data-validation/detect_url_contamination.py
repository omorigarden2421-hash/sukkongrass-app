# -*- coding: utf-8 -*-
"""参照URL/理由列に別植物のURLが混入している行を検出（同義学名・解析ノイズ除外）。"""
import openpyxl, re
wb=openpyxl.load_workbook('data-validation/納品_最終版/宿根草マスタ_検証済み_最終版.xlsx')
ws=wb.active
data=[[ws.cell(row=r,column=c).value for c in range(1,71)] for r in range(2,ws.max_row+1)]
def C(row,i): return str(row[i-1] or '').strip()
def genus(sci):
    m=re.match(r'([A-Z][a-z]+)', sci.strip()); return m.group(1).lower() if m else ''
SYN=[{'cimicifuga','actaea'},{'gaura','oenothera'},{'sedum','hylotelephium'},{'andropogon','schizachyrium'},
 {'polygonum','bistorta','persicaria'},{'chrysanthemum','nipponanthemum','dendranthema','ajania','leucanthemum','nipponicum'},
 {'ranunculus','ficaria'},{'eupatorium','conoclinium','ageratina','eutrochium'},
 {'aster','doellingeria','eurybia','symphyotrichum'},{'stachys','betony'},{'uncinia','carex'}]
def same(a,b): return a==b or any(a in s and b in s for s in SYN)
ART={'i','for','search','mouse','red','anacis','the','details','plant','plants','www','go','guides','pf','profile','wiki','result','uid','catalog','user'}
GEN=re.compile(r'(?:/plants?/|/plant/|/pflanzen/|LatinName=|name=)(?:\d+/)?([a-z]{4,})[-+ ]([a-z]+)')
hits=[]
for row in data:
    g=genus(C(row,4))
    if not g: continue
    for ci in [44,52,58,59,60,43,18,28]:
        for u in re.findall(r'https?://[^\s、,）)"]+', C(row,ci)):
            m=GEN.search(u.lower())
            if m and m.group(1) not in ART and not same(g, m.group(1)):
                hits.append((C(row,2), C(row,3).replace('\n',' '), ws.cell(row=1,column=ci).value, g, m.group(1), u))
seen=set()
for h in hits:
    k=(h[0],h[2],h[4])
    if k in seen: continue
    seen.add(k)
    print(f"{h[0]}番 {h[1][:16]:16s} [{h[2]}] {h[3]}→{h[4]}")
print('計', len(seen), '件')
