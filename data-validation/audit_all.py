# -*- coding: utf-8 -*-
"""全列横断の整合監査。数値異常・派生帯・タグ↔値・管理/生態の論理矛盾を検出し分類集計。"""
import openpyxl, re
from collections import Counter
wb=openpyxl.load_workbook('data-validation/納品_最終版/宿根草マスタ_検証済み_最終版.xlsx')
ws=wb.active
data=[[ws.cell(row=r,column=c).value for c in range(1,71)] for r in range(2,ws.max_row+1)]
def C(row,i): return str(row[i-1] or '').strip()
def nums(v): return [float(x) for x in re.findall(r'\d+\.?\d*', str(v))]
def mx(v): n=nums(v); return max(n) if n else None
def mn(v): n=nums(v); return min(n) if n else None

findings=[]
def F(cat,row,detail): findings.append((cat, C(row,2), C(row,3).replace('\n',' ')[:20], detail))

for row in data:
    nm=C(row,3)
    # ---- 数値サニティ ----
    w=mx(C(row,7)); rt=mx(C(row,8))
    if w and rt and w>rt: F('価格:卸>小売',row,f'卸{C(row,7)} > 小売{C(row,8)}')
    z1=mn(C(row,20)); z2=mx(C(row,21))
    if z1 and z2 and z1>z2: F('USDAゾーン下限>上限',row,f'{C(row,20)}>{C(row,21)}')
    for ci,cn in [(12,'高さ'),(13,'広がり')]:
        n=nums(C(row,ci))
        if len(n)>=2 and n[0]>n[-1] and '(' not in C(row,ci) and '（' not in C(row,ci):
            F(f'{cn}:min>max',row,C(row,ci))
    # ---- 高さ↔景観(61) ----
    H=mx(C(row,12)); land=C(row,61)
    if H is not None and land:
        band='前景' if H<=30 else ('中景' if H<=80 else '後景')
        if land in ('前景','中景','後景') and land!=band:
            F('高さ↔景観帯',row,f'高さ{C(row,12)}(={band}相当) だが景観={land}')
    # ---- 耐寒性(19)↔USDA下限(20) ----
    kan=C(row,19); zl=mn(C(row,20))
    if zl is not None:
        if kan=='強' and zl>=7: F('耐寒↔USDA',row,f'耐寒=強 だが下限ゾーン{C(row,20)}(高い=寒さ弱)')
        if kan=='弱' and zl<=4: F('耐寒↔USDA',row,f'耐寒=弱 だが下限ゾーン{C(row,20)}(低い=寒さ強)')
    # ---- タグ(62)↔値 ----
    func=C(row,62)
    if '蜜源' in func and C(row,55) not in ('◎','○'): F('タグ:蜜源↔蜜源価値',row,f'蜜源タグ だが蜜源価値={C(row,55)}')
    if '切り花' in func and C(row,37) not in ('◎','○'): F('タグ:切り花↔切り花適性',row,f'切り花タグ だが適性={C(row,37)}')
    if 'カラーリーフ' in func and C(row,40).startswith('①'): F('タグ:カラーリーフ↔葉色',row,f'カラーリーフタグ だが葉色={C(row,40)}')
    if ('乾燥地向き' in func or 'ロックガーデン' in func) and C(row,24)=='弱': F('タグ:乾燥向き↔乾燥耐性',row,f'乾燥/ロックタグ だが乾燥耐性=弱')
    if '日陰庭向き' in func and '陰' not in C(row,14): F('タグ:日陰↔日照',row,f'日陰タグ だが日照={C(row,14)}')
    # ---- デザインタグ(63)↔葉色(40) ----
    dz=C(row,63); leaf=C(row,40)
    if 'シルバー' in dz and 'シルバー' not in leaf and '④' not in leaf: F('タグ:シルバー↔葉色',row,f'シルバータグ だが葉色={leaf}')
    if '斑入り' in dz and '斑入り' not in leaf and '⑦' not in leaf: F('タグ:斑入り↔葉色',row,f'斑入りタグ だが葉色={leaf}')
    # ---- 生態整合 ----
    dry=C(row,24); soil=C(row,23)
    if dry=='強' and ('湿潤' in soil or '湿地' in soil): F('乾燥耐性強↔土壌湿潤',row,f'乾燥耐性強 だが土壌={soil}')
    if dry=='弱' and ('乾' in soil and '湿' not in soil): F('乾燥耐性弱↔土壌乾燥',row,f'乾燥耐性弱 だが土壌={soil}')
    if C(row,14)=='日陰' and C(row,53)=='◎': F('日陰↔西日耐性◎',row,f'日照=日陰 だが西日耐性=◎')
    # ---- 管理整合 ----
    mtype=C(row,45); habit=C(row,35); under=mn(C(row,50)); star=C(row,51)
    if mtype.startswith('E') and (under is None or under<=1) and '地下茎' not in habit and 'ランナー' not in habit and C(row,34) in ('低','なし',''):
        F('管理E↔非拡大',row,f'E隔離推奨 だが拡大性乏しい(習性={habit},地下茎Lv={C(row,50)})')
    if star=='★' and mtype.startswith('E'): F('管理手間★↔E型',row,f'管理手間★(易) だがE隔離推奨型')

cats=Counter(f[0] for f in findings)
print('=== 横断整合監査 検出サマリ ===')
for c,n in cats.most_common(): print(f'  {n:4d}  {c}')
print('  合計', len(findings))
print()
for cat in cats:
    ex=[f for f in findings if f[0]==cat][:4]
    print(f'--- {cat} ({cats[cat]}件) ---')
    for f in ex: print(f'   {f[1]}番 {f[2]:20s} {f[3]}')
