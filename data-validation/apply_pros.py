# -*- coding: utf-8 -*-
"""51 管理手間★ / 65 おすすめ_育てる環境 / 66 おすすめ_庭づくり を再導出。
   引数 --apply で両xlsxへ書き込み、無指定はプレビュー(件数+サンプル)。"""
import openpyxl, re, csv, sys
from collections import Counter

APPLY = '--apply' in sys.argv
FILES = [
    '納品_最終版/宿根草マスタ_検証済み_最終版.xlsx',
    '納品_最終版/宿根草マスタ_調査完了項目_赤塗り_最終版.xlsx',
]
NL = '\n'

def maxh(v):
    n=[int(x) for x in re.findall(r'\d+', str(v))] if v is not None else []
    return max(n) if n else None
def zi(v):
    try: return int(v)
    except: return None
def lines(v):
    return [x.strip() for x in str(v or '').split(NL) if x.strip()]

# ---- 65 育てる環境 ----
def derive65(c):
    sun=str(c[14] or '').strip(); low=zi(c[20]); heat=str(c[15] or '').strip()
    mure=str(c[16] or '').strip(); nishi=str(c[53] or '').strip(); dry=str(c[24] or '').strip()
    out=[]
    # 日照
    out.append({'日向':'日当たりの良い場所で庭づくりをされている方',
                '日向～半日陰':'日向〜半日陰まで幅広い環境で庭づくりをされている方',
                '半日陰':'半日陰の場所を活かして庭づくりをされている方',
                '明るい日陰':'半日陰〜日陰の難しい場所を活かしたい方'}.get(sun))
    # 耐寒
    if low is not None:
        if low<=5: out.append('寒冷地（北海道・東北）で越冬させたい方')
        elif low<=7: out.append('関東以北の庭で使いたい方')
    # 夏
    if heat=='強' and mure=='強': out.append('高温多湿の夏でも安心して育てたい方')
    elif heat=='弱' or mure=='弱': out.append('夏の高温多湿が少ない涼しい地域で庭づくりをされている方')
    # 西日
    if nishi=='◎': out.append('西向き・西日の強い場所に植えたい方')
    elif nishi=='×': out.append('西日を避けられる東〜南向きの場所で育てたい方')
    # 乾燥
    if dry=='強': out.append('水やりの手間を減らしたい方・乾きやすい場所に植えたい方')
    return NL.join([x for x in out if x])

# ---- 66 庭づくり ----
SUBJ66=['落ち着いた自然な雰囲気の庭をつくりたい方','一株で主役になる植物を探している方',
        'ボーダーガーデンをつくりたい方','まとめて群植して景観をつくりたい方',
        'コンテナや鉢植えでも育てたい方','モダン・ナチュラルなデザインの庭をつくりたい方']
def derive66(c):
    land=str(c[61] or '').strip(); cut=str(c[37] or '').strip(); mel=str(c[55] or '').strip()
    hab=str(c[35] or '').strip(); wet=str(c[23] or '').strip(); leaf=str(c[40] or '').strip()
    sil=str(c[57] or '').strip(); t62=lines_join(c[62]); t63=lines_join(c[63])
    cur=lines(c[66])
    out=[]
    out.append({'前景':'花壇や植栽の前列（手前側）を華やかに仕上げたい方',
                '中景':'植栽の中段レイヤーに動きと奥行きを加えたい方',
                '後景':'植栽の後方に高さと存在感を出したい方'}.get(land))
    if cut in ('◎','○'): out.append('切り花として室内でも楽しみたい方')
    if mel in ('◎','○'): out.append('蝶や蜂を呼び、生き物と共存する庭をつくりたい方')
    if hab in ('マット型','ランナー型'):
        out.append('地面を覆うグラウンドカバーとして使いたい方')
        out.append('斜面や傾斜地を緑化したい方')
    if '乾燥地向き' in t62: out.append('乾燥しがちな場所を活かしたい方')
    if wet=='湿潤寄り': out.append('水はけの悪い湿りがちな場所を活かしたい方')
    if leaf.startswith('⑦'): out.append('斑入り葉で明るいアクセントをつくりたい方')
    if leaf.startswith('④'): out.append('シルバーリーフで明るいコントラストをつくりたい方')
    if leaf.startswith('⑤'): out.append('黒葉でシックなアクセントを加えたい方')
    if sil=='グラス': out.append('グラス類を使ってナチュラルな雰囲気を出したい方')
    if 'ドライフラワー' in t62: out.append('ドライフラワーとしても活用したい方')
    if '北海道向き強健' in t63: out.append('北海道でも力強く育つ植物を求めている方')
    if 'モダン庭向き' in t63 and 'モダン・ナチュラルなデザインの庭をつくりたい方' not in cur:
        pass
    # 主観フレーズは現状維持（現行セルにあれば温存）
    subj=[s for s in SUBJ66 if s in cur]
    return NL.join([x for x in out if x] + subj)

def lines_join(v):
    return str(v or '')

# ---- 51 管理手間★（案A: 整合修正＋空欄補完） ----
BASE={'A 放任型':'★','B 花後刈込型':'★★','C 夏前切戻型':'★★★','D 毎年株更新型':'★★★★','E 隔離推奨型':'★★★★'}
def derive51(c):
    cur=str(c[51] or '').strip(); mtype=str(c[45] or '').strip()
    if cur in ('','None'):
        return BASE.get(mtype,'★★'), 'fill'
    # 整合修正: D/E型で★・★★は過小 → ★★★へ
    if mtype in ('D 毎年株更新型','E 隔離推奨型') and cur in ('★','★★'):
        return '★★★', 'fix'
    return cur, 'keep'

COLS=[14,15,16,20,23,24,35,37,40,45,51,53,55,57,61,62,63,65,66,3]

wb0=openpyxl.load_workbook(FILES[0]); ws0=wb0.active
new={}; diff={51:0,65:0,66:0}; fill51=0; fix51=0; samp={51:[],65:[],66:[]}
log=[]
for r in range(2, ws0.max_row+1):
    c={i:ws0.cell(row=r,column=i).value for i in COLS}
    name=str(c[3] or '').replace(NL,' ')
    n65=derive65(c); n66=derive66(c); n51,kind=derive51(c)
    new[r]=(n51,n65,n66)
    if kind=='fill': fill51+=1
    if kind=='fix': fix51+=1
    for col,nv,ov in [(51,n51,str(c[51] or '').strip()),(65,n65,str(c[65] or '').strip()),(66,n66,str(c[66] or '').strip())]:
        if nv!=ov:
            diff[col]+=1
            if len(samp[col])<4: samp[col].append((name[:22],ov,nv))
            log.append({'excel_row':r,'plant':name,'列':col,'旧値':ov,'新値':nv})

print(f'変更件数: 51管理手間={diff[51]}(fill{fill51}/fix{fix51})  65育てる環境={diff[65]}  66庭づくり={diff[66]}')
for col in (65,66,51):
    print(f'\n=== col{col} サンプル ===')
    for nm,o,n in samp[col]:
        print(f'[{nm}]')
        print(f'  旧: {o[:120]!r}')
        print(f'  新: {n[:120]!r}')

if APPLY:
    with open('管理手間_おすすめ文_変更ログ.csv','w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=['excel_row','plant','列','旧値','新値']); w.writeheader(); w.writerows(log)
    for path in FILES:
        wb=openpyxl.load_workbook(path); ws=wb.active
        for r in range(2, ws.max_row+1):
            c={i:ws.cell(row=r,column=i).value for i in COLS}
            n65=derive65(c); n66=derive66(c); n51,_=derive51(c)
            ws.cell(row=r,column=51).value=n51
            ws.cell(row=r,column=65).value=n65
            ws.cell(row=r,column=66).value=n66
        wb.save(path)
        print('適用完了',path)
    print('変更ログ', len(log),'行')
