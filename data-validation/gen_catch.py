# -*- coding: utf-8 -*-
"""一言キャッチ(64)を検証済み事実からルール生成。--apply で両xlsx反映、無指定はサンプル。"""
import openpyxl, re, csv, sys, random

APPLY='--apply' in sys.argv
FILES=['納品_最終版/宿根草マスタ_検証済み_最終版.xlsx','納品_最終版/宿根草マスタ_調査完了項目_赤塗り_最終版.xlsx']

def zi(v):
    try: return int(v)
    except: return None

def size_word(land):
    return {'前景':'コンパクトな','中景':'中型の','後景':'大型の'}.get(land,'')

def env_hook(dry, heat, mure, low, sun):
    if dry=='強': return '乾燥に強く育てやすい'
    if heat=='強' and mure=='強': return '暑さ・寒さに強い'
    if low is not None and low<=5: return '耐寒性に優れた'
    if '日陰' in sun: return '半日陰でも育つ'
    return '育てやすい'

def s1(c):
    land=str(c[61] or '').strip(); leaf=str(c[40] or '').strip(); sil=str(c[57] or '').strip()
    hab=str(c[35] or '').strip(); dry=str(c[24] or '').strip(); heat=str(c[15] or '').strip()
    mure=str(c[16] or '').strip(); low=zi(c[20]); sun=str(c[14] or '').strip()
    sz=size_word(land)
    # 1) グラス
    if sil=='グラス':
        return f'風にそよぐ葉が魅力の{sz}グラス'
    # 2) カラーリーフ
    if leaf[:1] in ('②','③','④','⑤','⑥','⑦'):
        lead={'④':'銀葉が美しい','⑦':'斑入り葉が映える','⑤':'シックな葉色が際立つ',
              '②':'明るい葉色が引き立つ','③':'涼しげな葉色が魅力の','⑥':'深い葉色が印象的な'}.get(leaf[:1],'彩り豊かな')
        return f'{lead}{sz}カラーリーフ'
    hook=env_hook(dry,heat,mure,low,sun)
    # 3) グラウンドカバー
    if hab in ('マット型','ランナー型'):
        return f'{hook}、よく広がる{sz}グラウンドカバー'
    # 4) 後景
    if land=='後景':
        return f'{hook}{sz}後景植物'
    # 5) 一般
    return f'{hook}{sz}品種'

def clean_term(s):
    # 括弧・全角空白・読点以降を落として主要語のみ
    s=re.split(r'[（(　,、/]', str(s or ''))[0].strip()
    return s

def s2(c):
    bloom=clean_term(c[10]); color=clean_term(c[11]); cut=str(c[37] or '').strip()
    sil=str(c[57] or '').strip()
    no_showy = (bloom in ('','ー','—','-','なし')) or (color in ('','目立たない','なし','緑','緑色'))
    if no_showy or sil=='グラス':
        if sil=='グラス': return '風になびく葉姿でナチュラルな庭に動きを添えます'
        return '花は目立たず、葉姿を楽しむ品種です'
    tail='、切り花にも活躍します' if cut=='◎' else 'ます'
    return f'{bloom}に{color}の花が咲き{tail}'

def s3(c, used):
    mel=str(c[55] or '').strip(); heat=str(c[15] or '').strip(); mure=str(c[16] or '').strip()
    sun=str(c[14] or '').strip()
    if heat=='弱' or mure=='弱': return '夏は涼しい場所が安心です'
    if mel=='◎': return '蝶や蜂も集まります'
    if '日陰' in sun and 'グラウンドカバー' not in used and '後景' not in used and 'カラーリーフ' not in used:
        return '半日陰でもよく育ちます'
    return ''

def gen(c):
    a=s1(c); b=s2(c); cc=s3(c,a)
    parts=[a,b]+([cc] if cc else [])
    return '。'.join(parts)+'。'

COLS=[10,11,14,15,16,20,24,35,37,40,55,57,61,64,3]
wb=openpyxl.load_workbook(FILES[0], read_only=not APPLY)
ws=wb.active

if not APPLY:
    data=list(ws.iter_rows(min_row=2,values_only=True))
    def C(r,i): return r[i-1]
    picks=[0,1,2,3,5,6,13,16,19]  # 多様な型
    # グラス/GC/後景/foliageも拾う
    for idx,r in enumerate(data):
        if len(picks)>=18: break
        leaf=str(C(r,40) or ''); sil=str(C(r,57) or ''); land=str(C(r,61) or '')
        if (sil=='グラス' or leaf.startswith('④') or land=='後景') and idx not in picks:
            picks.append(idx)
    for idx in picks[:18]:
        r=data[idx]; c={i:C(r,i) for i in COLS}
        nm=str(c[3] or '').replace(chr(10),' ')[:22]
        old=str(c[64] or '')[:110]
        print('[%s] 景観=%s 葉=%s 花=%s/%s' % (nm,str(c[61]),str(c[40])[:5],str(c[10]),str(c[11])))
        print('  旧: '+old)
        print('  新: '+gen(c))
        print()
else:
    log=[]
    for path in FILES:
        wb=openpyxl.load_workbook(path); ws=wb.active
        for r in range(2, ws.max_row+1):
            c={i:ws.cell(row=r,column=i).value for i in COLS}
            old=str(c[64] or '').strip(); new=gen(c)
            ws.cell(row=r,column=64).value=new
            if path==FILES[0] and new!=old:
                log.append({'excel_row':r,'plant':str(c[3] or '').replace('\n',' '),'旧値':old,'新値':new})
        wb.save(path); print('適用完了',path)
    with open('一言キャッチ_再生成ログ.csv','w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=['excel_row','plant','旧値','新値']); w.writeheader(); w.writerows(log)
    print('変更ログ',len(log),'行')
