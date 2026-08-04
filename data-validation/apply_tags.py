# -*- coding: utf-8 -*-
"""タグ列(38/61/62/63)を確定ルールで再導出し、マスタ2ファイルに適用＋変更ログ出力"""
import openpyxl, re, csv

FILES = [
    '納品_最終版/宿根草マスタ_検証済み_最終版.xlsx',
    '納品_最終版/宿根草マスタ_調査完了項目_赤塗り_最終版.xlsx',
]

def maxh(v):
    if v is None: return None
    n = [int(x) for x in re.findall(r'\d+', str(v))]
    return max(n) if n else None

def toks(v):
    s = '' if v is None else str(v).strip()
    return [t.strip() for t in re.split(r'[、,／/・\n]+', s) if t.strip()]

KEEP62 = ['ナチュラルガーデン向き', '強健（放置向き）', 'ドライフラワー', 'リーフ観賞']
ORD62  = ['蜜源', '切り花', '日陰庭向き', 'ロックガーデン向き', '乾燥地向き', 'カラーリーフ'] + KEEP62
KEEP63 = ['北海道向き強健', 'モダン庭向き']
ORD63  = ['シルバーリーフ', '斑入り葉', '黒葉アクセント', 'グラス調'] + KEEP63

COLS = {12:'h',14:'sun',23:'wet',24:'dry',35:'hab',37:'cut',40:'leaf',55:'mel',57:'sil',
        38:'c38',61:'c61',62:'c62',63:'c63',3:'name'}

def derive(cell):
    h=maxh(cell[12]); sun=str(cell[14] or '').strip(); dry=str(cell[24] or '').strip()
    wet=str(cell[23] or '').strip(); hab=str(cell[35] or '').strip(); cut=str(cell[37] or '').strip()
    leaf=str(cell[40] or '').strip(); mel=str(cell[55] or '').strip(); sil=str(cell[57] or '').strip()
    cur38,cur61,cur62,cur63=cell[38],cell[61],cell[62],cell[63]

    if h is None: n61=str(cur61 or '').strip()
    elif h<=30: n61='前景'
    elif h<=80: n61='中景'
    else: n61='後景'

    base={'前景':'前景・縁取り','中景':'中景構成','後景':'後景・背景'}.get(n61)
    gc = hab in ('マット型','ランナー型')
    n38 = str(cur38 or '').strip() if base is None else base+('、グラウンドカバー' if gc else '')

    s=set()
    if mel in ('◎','○'): s.add('蜜源')
    if cut in ('◎','○'): s.add('切り花')
    if '日陰' in sun: s.add('日陰庭向き')
    if dry=='強' or wet in ('排水必須','乾燥寄り'): s.add('ロックガーデン向き'); s.add('乾燥地向き')
    if leaf[:1] in ('②','③','④','⑤','⑥','⑦'): s.add('カラーリーフ')
    for t in KEEP62:
        if t in toks(cur62): s.add(t)
    n62='、'.join(t for t in ORD62 if t in s)

    d=set()
    if leaf.startswith('④'): d.add('シルバーリーフ')
    if leaf.startswith('⑦'): d.add('斑入り葉')
    if leaf.startswith('⑤'): d.add('黒葉アクセント')
    if sil=='グラス': d.add('グラス調')
    for t in KEEP63:
        if t in toks(cur63): d.add(t)
    n63='、'.join(t for t in ORD63 if t in d)
    return n38,n61,n62,n63

def norm(v): return str(v or '').strip()

# ログは検証済み版基準で1回生成
wb0=openpyxl.load_workbook(FILES[0]); ws0=wb0.active
log=[]
for r in range(2, ws0.max_row+1):
    cell={i:ws0.cell(row=r,column=i).value for i in COLS}
    name=str(cell[3] or '').replace('\n',' ')
    news=derive(cell)
    for col,new in zip((38,61,62,63),news):
        old=norm(cell[col])
        if new!=old:
            log.append({'excel_row':r,'plant':name,'列':col,
                        '列名':{38:'用途区分',61:'景観構成用途タグ',62:'機能用途タグ',63:'デザイン系タグ'}[col],
                        '旧値':old,'新値':new})
with open('タグ再導出_変更ログ.csv','w',encoding='utf-8-sig',newline='') as f:
    w=csv.DictWriter(f,fieldnames=['excel_row','plant','列','列名','旧値','新値']); w.writeheader(); w.writerows(log)
from collections import Counter
print('変更ログ', len(log),'行 ', dict(Counter(x['列名'] for x in log)))

for path in FILES:
    wb=openpyxl.load_workbook(path); ws=wb.active
    n=0
    for r in range(2, ws.max_row+1):
        cell={i:ws.cell(row=r,column=i).value for i in COLS}
        n38,n61,n62,n63=derive(cell)
        ws.cell(row=r,column=38).value=n38
        ws.cell(row=r,column=61).value=n61
        ws.cell(row=r,column=62).value=n62
        ws.cell(row=r,column=63).value=n63
        n+=1
    wb.save(path)
    print('適用完了',path,n,'行')
