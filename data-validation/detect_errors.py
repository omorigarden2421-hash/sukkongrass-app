# -*- coding: utf-8 -*-
"""マスタ誤り検出網: 定型注記スタンプの棚卸し＋構造化列(検証済)との矛盾検出＋属の生態不一致。
   出力: マスタ誤り疑いリスト.xlsx （疑いリスト / スタンプ棚卸し / 検出ロジック）"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

wb=openpyxl.load_workbook('data-validation/納品_最終版/宿根草マスタ_検証済み_最終版.xlsx')
ws=wb.active
hdr=[str(c.value or '') for c in ws[1]]
data=[[ws.cell(row=r,column=c).value for c in range(1,len(hdr)+1)] for r in range(2,ws.max_row+1)]
def C(row,i): return str(row[i-1] or '').strip()

NOTE=17          # 耐暑性に関する注記
HEAT=15          # 耐暑性(検証済)
DRY=24           # 乾燥耐性(検証済)
SUN=14           # 日照条件
ORIG=27          # 原産地
KAKONE=44        # 耐潮性_根拠
SHIO=42          # 耐潮性

SUCC_GENUS=['セダム','センペル','デロスペルマ','エケベリア','グラプト','オロスタキス','ロディオラ','セdum']
note_counts=Counter(C(row,NOTE) for row in data if C(row,NOTE))

def is_stamp(t): return t and note_counts.get(t,0)>=3

findings=[]  # (優先, 商品番号, 名, 学名, 原産地, 対象列, 現値, 検出理由, 波及行数)
def add(pri,row,col_i,reason):
    findings.append((pri, C(row,2), C(row,3).replace('\n',' '), C(row,4), C(row,ORIG),
                     hdr[col_i-1], C(row,col_i)[:70], reason, note_counts.get(C(row,col_i),1)))

for row in data:
    note=C(row,NOTE); heat=C(row,HEAT); dry=C(row,DRY); sun=C(row,SUN); nm=C(row,3)
    genus_succ=any(g in nm for g in SUCC_GENUS)
    # A: 多肉スタンプ×非多肉属（誤スタンプ濃厚）
    if ('多肉' in note or '乾燥地原産' in note) and not genus_succ:
        add('P1',row,NOTE,'非多肉属に「多肉/乾燥地原産」注記（属の生態不一致）')
    # B3: 注記=乾燥好き だが 乾燥耐性=弱（検証済列と矛盾）
    if ('多肉' in note or '乾燥を好む' in note or '乾燥地原産' in note) and dry=='弱':
        add('P1',row,NOTE,'注記は乾燥好きだが乾燥耐性(検証済)=弱と矛盾')
    # B5: 注記=湿地/湿潤/森林下草 だが 乾燥耐性=強
    if any(k in note for k in ['湿地','湿潤土壌','森林下草','高湿度環境に適応']) and dry=='強':
        add('P2',row,NOTE,'注記は湿潤好きだが乾燥耐性(検証済)=強と矛盾')
    # B1: 注記=冷涼地/夏越し困難 だが 耐暑性=強（強い主張のみ）
    if any(k in note for k in ['冷涼地原産','夏越しが困難','夏越し困難','高温多湿で枯れ']) and heat=='強':
        add('P1',row,NOTE,'注記は暑さに弱いが耐暑性(検証済)=強と矛盾')
    # B2: 注記=高温多湿に強い という強い断定 だが 耐暑性=弱（「夏越し可能」等の条件付きは除外）
    if any(k in note for k in ['高温多湿耐性あり','夏越し安定','高湿度でも安定','高温多湿耐性']) and heat=='弱':
        add('P1',row,NOTE,'注記は高温多湿に強いと断定だが耐暑性(検証済)=弱と矛盾')
    # B4: 注記=森林下草/半日陰/日陰 だが 日照=日向のみ
    if any(k in note for k in ['森林下草','半日陰','日陰','木漏れ日']) and sun=='日向':
        add('P3',row,NOTE,'注記は日陰性だが日照条件=日向のみ（要確認）')
    # C: 耐潮性根拠に他分類群/多肉混入
    kk=C(row,KAKONE)
    if ('多肉' in kk or 'Mangave' in kk or 'アガベ' in kk) and not genus_succ and 'ハマ' not in nm and 'ギク' not in nm:
        add('P1',row,KAKONE,'耐潮性根拠に別分類群/多肉の記述が混入の疑い')

# 重複finding除去（同一行・同一列・同一理由）
seen=set(); uniq=[]
for f in findings:
    k=(f[1],f[5],f[7])
    if k not in seen: seen.add(k); uniq.append(f)
pri_order={'P1':0,'P2':1,'P3':2}
uniq.sort(key=lambda f:(pri_order[f[0]], -f[8]))

n1=sum(1 for f in uniq if f[0]=='P1'); n2=sum(1 for f in uniq if f[0]=='P2'); n3=sum(1 for f in uniq if f[0]=='P3')
print('検出: 総 %d件  P1=%d P2=%d P3=%d' % (len(uniq),n1,n2,n3))

# ===== Excel出力 =====
out=openpyxl.Workbook()
s1=out.active; s1.title='疑いリスト'
cols=['優先度','商品番号','植物名','学名','原産地','対象列','現在の記述(抜粋)','検出理由','同一スタンプ波及行数']
s1.append(cols)
fillP1=PatternFill('solid',fgColor='FFC7CE'); fillP2=PatternFill('solid',fgColor='FFEB9C'); fillP3=PatternFill('solid',fgColor='DDEBF7')
for f in uniq:
    s1.append(list(f))
    c=s1.cell(row=s1.max_row,column=1)
    c.fill={'P1':fillP1,'P2':fillP2,'P3':fillP3}[f[0]]
for j,w in enumerate([8,10,26,34,26,20,50,44,12],1):
    s1.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
for cell in s1[1]: cell.font=Font(bold=True); cell.fill=PatternFill('solid',fgColor='4472C4'); cell.font=Font(bold=True,color='FFFFFF')
s1.freeze_panes='A2'

# スタンプ棚卸し
s2=out.create_sheet('スタンプ棚卸し')
s2.append(['注記(定型文)','使用行数','強い生態主張'])
risky=['乾燥地','多肉','冷涼地','森林下草','湿地','高山','砂漠','海岸','日陰']
for txt,n in note_counts.most_common():
    if n>=3:
        s2.append([txt, n, '●' if any(k in txt for k in risky) else ''])
for j,w in enumerate([90,10,12],1):
    s2.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
for cell in s2[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='4472C4')
s2.freeze_panes='A2'

# ロジック説明
s3=out.create_sheet('検出ロジック')
logic=[['検出器','内容','優先度'],
['A 属の生態不一致','多肉属でないのに「多肉/乾燥地原産」注記 → 誤スタンプ濃厚','P1'],
['B1 冷涼×耐暑強','注記は暑さに弱いが検証済の耐暑性=強と矛盾','P1'],
['B2 高温OK×耐暑弱','注記は暑さに強いが検証済の耐暑性=弱と矛盾','P1'],
['B3 乾燥好き×乾燥弱','注記は乾燥好きだが検証済の乾燥耐性=弱と矛盾','P1'],
['B5 湿潤好き×乾燥強','注記は湿潤好きだが検証済の乾燥耐性=強と矛盾','P2'],
['B4 日陰×日向','注記は日陰性だが日照条件=日向のみ（日照は未検証のため要確認）','P3'],
['C 根拠の分類群混入','耐潮性根拠に別分類群/多肉の記述が混入','P1'],
['','',''],
['方針','構造化列(耐暑性/乾燥耐性)は1品種ずつ検証済のため、注記と矛盾する場合は注記側が誤りの可能性が高い',''],
['再発防止','注記は保存前に必ず同行の原産地・構造化列と照合。属スタンプは1品種確認後のみ',''] ]
for r in logic: s3.append(r)
for j,w in enumerate([20,70,10],1):
    s3.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
for cell in s3[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='4472C4')

out.save('data-validation/マスタ誤り疑いリスト.xlsx')
print('出力: data-validation/マスタ誤り疑いリスト.xlsx')
print()
print('=== P1（誤りほぼ確定）一覧 ===')
for f in uniq:
    if f[0]=='P1':
        print(f'  {f[1]}番 {f[2][:20]:20s} [{f[5]}] {f[7]}')
